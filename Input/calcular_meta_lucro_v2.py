"""
GM PLANNING — CÁLCULO INCREMENTAL DA META DE LUCRO POR TERRITORY
=======================================================================
- Este script implementa a metodologia GM Planning definida e
validada para a distribuição da meta de Lucro da operação de Parcerias
entre as Zonas de Venda (TERRITORY), seguindo as regras estabelecidas.

- O objetivo de lucro é um input direto do usuário 
(meta_lucro_total em CONFIG, igual para todos os meses) e é repartido
entre os TERRITORY com base no comportamento histórico real de cada um.

ETAPA 0. MODO DE OPERAÇÃO - ARQUIVO ÚNICO E INCREMENTAL

      - O script mantém 1 arquivo de saída (arquivo_saida) que
      acumula a meta de todos os meses desde mes_inicio_historico
      até o "mês corrente".
      
      - "Mês corrente" é, por padrão, o mês atual do calendário no
      momento em que o script roda (pode ser forçado via
      config["mes_corrente"], útil para testes). 
      
      - A cada execução, o script:
          
        1. Lê o arquivo de saída existente (se já houver um) e
           identifica quais meses já estão calculados nele.
           
        2. Monta a lista de todos os meses de mes_inicio_historico até
           o mês corrente (sequência mensal cheia, mês a mês).
           
        3. Calcula a meta APENAS dos meses dessa lista que AINDA NÃO
           estão no arquivo -- nunca recalcula um mês que já foi
           gravado anteriormente, mesmo que o histórico em
           tb_resultado tenha mudado depois. Isso vale tanto para a
           primeira execução (que faz o backfill de todos os meses
           desde mes_inicio_historico) quanto para as seguintes (que,
           rodando uma vez por mês, normalmente só terão o novo "mês
           corrente" para calcular).
           
        4. Junta o que já existia com o que foi calculado agora e
           regrava o arquivo (mesmo conteúdo dos meses antigos, só
           acrescido das linhas novas).

ETAPA 1. Distribuição da meta de Lucro entre TERRITORY (por mês)

      - A meta de Lucro total (meta_lucro_total, a mesma para todo mês)
      é repartida entre os TERRITORY com base num "share" histórico de
      cada zona, calculado como uma média ponderada entre:
        • participação da zona no Faturamento real total (peso 70%)
        • participação da zona no número de PDVs total (peso 30%)
        
      - O share é calculado mês a mês e resumido pela MEDIANA de cada TERRITORY. 
      Os shares medianos são renormalizados para somar 100%, garantindo que
      a soma das metas de Lucro por zona feche exatamente meta_lucro_total do mês.

ETAPA 2 — Meta de Faturamento e Teto de Spread por TERRITORY
  
      - Como Lucro = Faturamento x (GM Partidor - Spread), a meta de
      Lucro de cada zona (Etapa 1.1) precisa ser "traduzida" em um
      Faturamento-alvo e um teto de bonificação (Spread) compatíveis.
      
      - Partimos da mediana histórica de Faturamento, GM Partidor e
      Spread de cada zona e movemos Faturamento e Spread (KPIs acionáveis) 
      em unidades de desvio-padrão histórico, mantendo o GM Partidor quase fixo na
      mediana. 
      
      - O deslocamento é limitado a +/- K_MAX desvios-padrão; se isso não bastar 
      para fechar a meta de Lucro, o resíduo é fechado ajustando o GM Partidor internamente.
      
      - O Spread resultante é reportado como TETO SPREAD: o nível máximo
      de bonificação que a zona pode praticar, dado o Faturamento-alvo
      e o mix (GM Partidor), para não comprometer a meta de Lucro.

ETAPA 3 — Exportação incremental

      - Geração/atualização de um único arquivo .xlsx com uma linha por
      TERRITORY x MÊS.
      
      - Colunas: MÊS; TERRITORY; META LUCRO;
      META FATURAMENTO; GM PARTIDOR; TETO SPREAD.

OUTROS - Regra de Eligibilidade Mínima

      Um TERRITORY só recebe meta calculada num determinado mês se
      tiver, em tb_resultado, pelo menos historico_minimo_meses
      (padrão: 3) meses de dado real anteriores àquele mês.
      
      Isso vale em dois níveis: 
          
    (a) No dataset como um todo (se não houver histórico geral
      suficiente antes de um mês da lista, esse mês é pulado com um
      aviso, sem interromper o cálculo dos demais);
    
    (b) Por TERRITORY (zonas novas, sem histórico suficiente ainda, ficam de
      fora do universo daquele mês até acumularem histórico).
      Como tb_resultado começa em outubro/2025, janeiro/2026 é o
      primeiro mês em que a condição (a) já é satisfeita (histórico:
      outubro, novembro e dezembro) -- por isso é o mês de início
      padrão (mes_inicio_historico).

As configurações ficam centralizadas no dicionário CONFIG abaixo.
"""

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =========================================================================
# CONFIGURAÇÃO
# =========================================================================
CONFIG = {
    # arquivo de entrada (única fonte de dados)
    "arquivo_resultado": "[input] tb_resultado.xlsx",

    # primeiro mês a ser calculado (o backfill inicial começa aqui)
    "mes_inicio_historico": "2026-01-01",

    # "mês corrente": até qual mês calcular. None = usa o mês atual do
    # calendário (data de hoje, no momento em que o script roda). Pode
    # ser fixado (ex. "2026-08-01") para testes/reprodutibilidade.
    "mes_corrente": None,

    # meta de lucro total (R$) -- a MESMA para todos os meses calculados
    "meta_lucro_total": 5_000_000.0,

    # histórico mínimo exigido (em meses) antes de cada mês calculado
    "historico_minimo_meses": 3,

    # pesos do blend de share (Etapa 1.1) -- devem somar 1.0
    "peso_faturamento": 0.70,
    "peso_pdv": 0.30,

    # tratamento de outliers e histórico curto (Etapa 1.2)
    "spread_cap_outlier": 0.30,     # teto de winsorização do Spread histórico
    "desvio_padrao_fallback": 0.15,  # desvio de referência (15% da mediana)
    "k_max_desvios": 2.0,           # limite de +/- desvios-padrão no ajuste

    # arquivo de saída ÚNICO e persistente (sem placeholder de mês --
    # acumula todos os meses ao longo do tempo)
    "arquivo_saida": "[output] tb_meta_sop.xlsx",
}

# nomes das colunas "bonitas" (arquivo final) <-> nomes internos (cálculo)
COLUNAS_SAIDA = ["MÊS", "TERRITORY", "META LUCRO", "META FATURAMENTO", "GM PARTIDOR", "TETO SPREAD"]
COLUNAS_INTERNAS = ["MES", "TERRITORY", "META_LUCRO", "META_FATURAMENTO", "GM_PARTIDOR", "TETO_SPREAD"]
MAPA_SAIDA_PARA_INTERNO = dict(zip(COLUNAS_SAIDA, COLUNAS_INTERNAS))


# =========================================================================
# ETAPA 0 — Carga de dados e resolução de datas
# =========================================================================
def carregar_dados(config):
    """Lê tb_resultado na íntegra (sem filtrar por mês)."""
    return pd.read_excel(config["arquivo_resultado"])


def resolver_mes_corrente(config):
    """
    Resolve o "mês corrente": usa config["mes_corrente"] se ele estiver
    definido, senão usa o mês atual do calendário (primeiro dia do mês
    de hoje).
    """
    if config.get("mes_corrente"):
        return pd.Timestamp(config["mes_corrente"]).to_period("M").to_timestamp()
    return pd.Timestamp.today().to_period("M").to_timestamp()


def carregar_resultado_existente(config):
    """
    Lê o arquivo de saída já existente, se houver, e devolve:
        df_existente     -- DataFrame com colunas internas
                             (MES, TERRITORY, META_LUCRO, ...)
        meses_calculados -- set de meses (Timestamp) já presentes nele
    Se o arquivo ainda não existir, devolve um DataFrame vazio (com as
    colunas certas) e um set vazio -- ou seja, a primeira execução faz
    o backfill completo desde mes_inicio_historico.
    """
    caminho = config["arquivo_saida"]
    if os.path.exists(caminho):
        bruto = pd.read_excel(caminho)
        bruto = bruto.rename(columns=MAPA_SAIDA_PARA_INTERNO)
        bruto["MES"] = pd.to_datetime(bruto["MES"])
        meses_calculados = set(bruto["MES"].unique())
        return bruto[COLUNAS_INTERNAS], meses_calculados
    return pd.DataFrame(columns=COLUNAS_INTERNAS), set()


# =========================================================================
# ETAPA 1.1 — Share de cada TERRITORY e distribuição da meta de Lucro
# (recebe res_hist já filtrado para o mês em cálculo -- agnóstica de mês)
# =========================================================================
def calcular_meta_lucro_por_territory(res_hist, config):
    """
    Implementa a Etapa 1.1: calcula, para cada TERRITORY, o share médio
    (mediana mensal) de participação no negócio, com base num blend
    ponderado entre share de Faturamento real (peso_faturamento) e share
    de número de PDVs (peso_pdv). Em seguida, distribui meta_lucro_total
    proporcionalmente a esse share, renormalizado para somar
    exatamente 100% entre os TERRITORY do universo final.

    O "universo final" é o conjunto de TERRITORY que aparecem em
    res_hist com pelo menos config["historico_minimo_meses"] meses de
    dado real -- única condição de elegibilidade.

    Retorna:
        meta_lucro   -- Series (index=TERRITORY) com a meta de Lucro (R$)
        universo     -- lista de TERRITORY no universo final
        diagnostico  -- dict com contagens para conferência/relatório
    """
    peso_fat = config["peso_faturamento"]
    peso_pdv = config["peso_pdv"]
    minimo = config["historico_minimo_meses"]

    # --- share mensal de cada TERRITORY sobre o total (PDVs e Faturamento) ---
    totais_mes = res_hist.groupby("PRICE_MONTH").agg(
        TOTAL_PDVS=("PDVs", "sum"),
        TOTAL_FAT=("FATURAMENTO", "sum"),
    )
    df = res_hist.merge(totais_mes, on="PRICE_MONTH", how="left")
    df["SHARE_PDV"] = df["PDVs"] / df["TOTAL_PDVS"]
    df["SHARE_FAT"] = df["FATURAMENTO"] / df["TOTAL_FAT"]
    df["SHARE_FINAL"] = peso_fat * df["SHARE_FAT"] + peso_pdv * df["SHARE_PDV"]

    # --- mediana do share mensal por TERRITORY (robusta a meses atípicos) ---
    share_mediano = df.groupby("TERRITORY")["SHARE_FINAL"].median()

    # --- nº de meses de histórico real que cada TERRITORY efetivamente tem ---
    meses_por_territory = res_hist.groupby("TERRITORY")["PRICE_MONTH"].nunique()

    # --- universo final: TERRITORY com pelo menos `minimo` meses de histórico ---
    territorios_com_algum_historico = set(meses_por_territory.index)
    universo = sorted(meses_por_territory[meses_por_territory >= minimo].index)

    diagnostico = {
        "territorios_com_algum_historico": len(territorios_com_algum_historico),
        "universo_final": len(universo),
        "excluidos_por_historico_abaixo_do_minimo": sorted(
            set(territorios_com_algum_historico) - set(universo)
        ),
    }

    if len(universo) == 0:
        raise ValueError(
            f"Nenhum TERRITORY tem pelo menos {minimo} meses de histórico real "
            "disponíveis para este mês."
        )

    # --- renormaliza o share para somar exatamente 100% no universo final ---
    share_universo = share_mediano.loc[universo]
    share_normalizado = share_universo / share_universo.sum()

    # --- distribui a meta de Lucro total proporcionalmente ao share ---
    meta_lucro = (share_normalizado * config["meta_lucro_total"]).rename("META_LUCRO")

    return meta_lucro, universo, diagnostico


# =========================================================================
# ETAPA 1.2 — Meta de Faturamento e Teto de Spread por TERRITORY
# (também agnóstica de mês -- nenhuma alteração necessária aqui)
# =========================================================================
def _resolver_k(fat_med, fat_std, gm_med, spread_med, spread_std, meta_lucro):
    """
    Resolve, para uma zona, o deslocamento 'k' (em desvios-padrão) a ser
    aplicado simultaneamente a Faturamento e Spread a partir de suas
    medianas, de forma que:

        Faturamento_novo = fat_med + k * fat_std
        Spread_novo      = spread_med - k * spread_std
        Lucro = Faturamento_novo * (gm_med - Spread_novo) = meta_lucro

    Isso equivale a uma equação do 2º grau em k. Entre as duas raízes
    possíveis, escolhe-se a de menor módulo (ajuste mais conservador).
    """
    a = fat_std * spread_std
    margem_base = gm_med - spread_med
    b = fat_med * spread_std + fat_std * margem_base
    c = fat_med * margem_base - meta_lucro

    if abs(a) < 1e-12:
        return 0.0 if abs(b) < 1e-12 else -c / b

    discriminante = b ** 2 - 4 * a * c
    if discriminante < 0:
        return -b / (2 * a)

    raiz = np.sqrt(discriminante)
    k1 = (-b + raiz) / (2 * a)
    k2 = (-b - raiz) / (2 * a)
    return k1 if abs(k1) <= abs(k2) else k2


def calcular_meta_faturamento_e_teto_spread(res_hist, universo, meta_lucro, config):
    """
    Implementa a Etapa 1.2 para cada TERRITORY do universo final:
      1. Mediana e desvio-padrão históricos de Faturamento, GM Partidor
         e Spread (Spread winsorizado em spread_cap_outlier antes do
         cálculo).
      2. Desvio-padrão de referência (desvio_padrao_fallback) para o
         caso raro de desvio zero.
      3. Resolve o deslocamento em desvios-padrão (k) que leva
         Faturamento e Spread, partindo de suas medianas, a produzirem
         a meta de Lucro da zona (GM Partidor quase fixo). k limitado
         a +/- k_max_desvios.
      4. Resíduo (se o cap de k não bastar) fechado ajustando o GM
         Partidor internamente.

    Retorna um DataFrame indexado por TERRITORY com META_FATURAMENTO,
    GM_PARTIDOR, TETO_SPREAD e colunas de diagnóstico.
    """
    spread_cap = config["spread_cap_outlier"]
    fallback_rel_std = config["desvio_padrao_fallback"]
    k_max = config["k_max_desvios"]

    base = res_hist[res_hist["TERRITORY"].isin(universo)].copy()
    base["SPREAD_CAPPED"] = base["SPREAD"].clip(lower=0, upper=spread_cap)

    stats = base.groupby("TERRITORY").agg(
        FAT_MED=("FATURAMENTO", "median"),
        FAT_STD=("FATURAMENTO", "std"),
        GM_MED=("GM_PARTIDOR", "median"),
        SPREAD_MED=("SPREAD_CAPPED", "median"),
        SPREAD_STD=("SPREAD_CAPPED", "std"),
        N_OBS=("FATURAMENTO", "count"),
    )

    stats["FAT_STD"] = np.where(
        stats["FAT_STD"].isna() | (stats["FAT_STD"] == 0),
        fallback_rel_std * stats["FAT_MED"],
        stats["FAT_STD"],
    )
    stats["SPREAD_STD"] = np.where(
        stats["SPREAD_STD"].isna() | (stats["SPREAD_STD"] == 0),
        fallback_rel_std * stats["SPREAD_MED"].clip(lower=0.01),
        stats["SPREAD_STD"],
    )

    stats = stats.join(meta_lucro)

    linhas = []
    for territory, linha in stats.iterrows():
        k = _resolver_k(
            linha["FAT_MED"], linha["FAT_STD"], linha["GM_MED"],
            linha["SPREAD_MED"], linha["SPREAD_STD"], linha["META_LUCRO"],
        )
        k_limitado = float(np.clip(k, -k_max, k_max))
        cap_atingido = not np.isclose(k, k_limitado, atol=1e-9)

        fat_novo = max(linha["FAT_MED"] + k_limitado * linha["FAT_STD"], 1.0)
        teto_spread = max(linha["SPREAD_MED"] - k_limitado * linha["SPREAD_STD"], 0.0)
        gm_usado = linha["GM_MED"]

        lucro_obtido = fat_novo * (gm_usado - teto_spread)
        residuo = linha["META_LUCRO"] - lucro_obtido
        gm_ajustado = False

        if abs(residuo) > 1.0 and fat_novo > 0:
            gm_necessario = np.clip(gm_usado + residuo / fat_novo, 0.0, 0.60)
            if not np.isclose(gm_necessario, gm_usado, atol=1e-6):
                gm_usado = gm_necessario
                gm_ajustado = True
                lucro_obtido = fat_novo * (gm_usado - teto_spread)

        linhas.append({
            "TERRITORY": territory,
            "META_LUCRO": round(linha["META_LUCRO"], 2),
            "META_FATURAMENTO": round(fat_novo, 2),
            "GM_PARTIDOR": round(gm_usado, 4),
            "TETO_SPREAD": round(teto_spread, 4),
            "_LUCRO_CONFERENCIA": round(lucro_obtido, 2),
            "_CAP_ATINGIDO": cap_atingido,
            "_GM_AJUSTADO_POR_RESIDUO": gm_ajustado,
        })

    return pd.DataFrame(linhas).set_index("TERRITORY")


# =========================================================================
# ETAPA 1.3 — Exportação (regrava o arquivo único com tudo: meses antigos
# preservados + meses novos recém-calculados)
# =========================================================================
def exportar_xlsx(resultado_completo, config):
    """
    Recebe o DataFrame combinado (meses antigos + meses novos, colunas
    internas COLUNAS_INTERNAS) e regrava o arquivo .xlsx único por
    completo, com formatação profissional. Como os valores dos meses
    antigos não são recalculados (só carregados de volta do próprio
    arquivo), o conteúdo deles permanece idêntico ao que já estava
    salvo -- só as linhas dos meses novos são, de fato, dado novo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Meta de Lucro"

    ws.append(COLUNAS_SAIDA)

    fonte_cabecalho = Font(name="Arial", bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill("solid", fgColor="1F3B73")
    for col_idx in range(1, len(COLUNAS_SAIDA) + 1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center")

    resultado_ordenado = resultado_completo.sort_values(["MES", "TERRITORY"])
    for _, linha in resultado_ordenado.iterrows():
        ws.append([
            pd.Timestamp(linha["MES"]).to_pydatetime(),
            linha["TERRITORY"],
            linha["META_LUCRO"],
            linha["META_FATURAMENTO"],
            linha["GM_PARTIDOR"],
            linha["TETO_SPREAD"],
        ])

    fonte_corpo = Font(name="Arial")
    n_linhas = len(resultado_ordenado)
    for row_idx in range(2, n_linhas + 2):
        ws.cell(row=row_idx, column=1).number_format = "MM/YYYY"
        ws.cell(row=row_idx, column=3).number_format = 'R$ #,##0'
        ws.cell(row=row_idx, column=4).number_format = 'R$ #,##0'
        ws.cell(row=row_idx, column=5).number_format = '0.0%'
        ws.cell(row=row_idx, column=6).number_format = '0.0%'
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).font = fonte_corpo

    larguras = [12, 16, 16, 18, 14, 14]
    for col_idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.freeze_panes = "A2"

    caminho_saida = config["arquivo_saida"]
    wb.save(caminho_saida)
    return caminho_saida


# =========================================================================
# MAIN — carrega o que já existe, calcula só os meses faltantes, salva
# =========================================================================
def main(config=CONFIG):
    res = carregar_dados(config)
    mes_inicio = pd.Timestamp(config["mes_inicio_historico"]).to_period("M").to_timestamp()
    mes_corrente = resolver_mes_corrente(config)

    todos_os_meses_alvo = pd.date_range(start=mes_inicio, end=mes_corrente, freq="MS")

    df_existente, meses_calculados = carregar_resultado_existente(config)
    meses_para_calcular = [m for m in todos_os_meses_alvo if m not in meses_calculados]

    print(f"Mês corrente: {mes_corrente:%Y-%m}")
    print(f"Meses já calculados no arquivo ({len(meses_calculados)}): "
          f"{sorted(m.strftime('%Y-%m') for m in meses_calculados)}")

    if not meses_para_calcular:
        print(f"\nNada a fazer: todos os meses de {mes_inicio:%Y-%m} até {mes_corrente:%Y-%m} "
              f"já estão em {config['arquivo_saida']}.")
        return df_existente

    print(f"Meses a calcular nesta execução: {[m.strftime('%Y-%m') for m in meses_para_calcular]}")

    novos_resultados = []
    for mes in meses_para_calcular:
        minimo = config["historico_minimo_meses"]
        hist_months = sorted(res.loc[res["PRICE_MONTH"] < mes, "PRICE_MONTH"].unique())
        res_hist = res[res["PRICE_MONTH"].isin(hist_months)].copy()

        print(f"\n=== Mês {mes:%Y-%m} (meta de lucro total: R$ {config['meta_lucro_total']:,.2f}) ===")

        if len(hist_months) < minimo:
            print(f"  IGNORADO: apenas {len(hist_months)} mês(es) de histórico disponível "
                  f"antes dele (mínimo exigido: {minimo}).")
            continue

        print(f"  Janela histórica: {[m.strftime('%Y-%m') for m in hist_months]}")

        try:
            meta_lucro, universo, diagnostico = calcular_meta_lucro_por_territory(res_hist, config)
        except ValueError as e:
            print(f"  IGNORADO: {e}")
            continue

        print(f"  Universo final (>= {minimo} meses de histórico): {diagnostico['universo_final']} "
              f"(excluídos por histórico insuficiente: {len(diagnostico['excluidos_por_historico_abaixo_do_minimo'])})")
        print(f"  Soma da meta de Lucro distribuída: R$ {meta_lucro.sum():,.2f}")

        resultado_mes = calcular_meta_faturamento_e_teto_spread(res_hist, universo, meta_lucro, config)
        print(f"  Territórios com ajuste limitado pelo cap (k)     : {int(resultado_mes['_CAP_ATINGIDO'].sum())}")
        print(f"  Territórios com GM Partidor ajustado por resíduo : {int(resultado_mes['_GM_AJUSTADO_POR_RESIDUO'].sum())}")

        resultado_mes = resultado_mes.reset_index()
        resultado_mes["MES"] = mes
        novos_resultados.append(resultado_mes[COLUNAS_INTERNAS])

    if not novos_resultados:
        print("\nNenhum mês pôde ser calculado nesta execução (ver avisos acima).")
        return df_existente

    resultado_completo = pd.concat([df_existente] + novos_resultados, ignore_index=True)
    caminho_saida = exportar_xlsx(resultado_completo, config)

    print(f"\nEtapa 1.3 — Arquivo atualizado em: {caminho_saida}")
    print(f"Total de meses no arquivo: {resultado_completo['MES'].nunique()} "
          f"| Total de linhas (TERRITORY x MÊS): {len(resultado_completo)}")

    return resultado_completo


if __name__ == "__main__":
    main()
